from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active


# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0,100)])

# col_B = ws["B"]
# print(col_B)
# for cell in col_B:
#     print(cell.value)
# col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] # 2번쨰 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end = " ")
#         # print(cell.coordinate, end = " ")
#         xy = coordinate_from_string(cell.coordinate)
#         print(xy, end = " ")
#     print()

# 전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)


#전체 columns
# print(tuple(ws.columns))
# for column in tuple(ws.columns):
#     print(column[0].value)


wb.save("sample.xlsx")

