from json import load
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 ( 국어) 영어 수학

# ws.move_range("B1:C11", rows = 0, cols = 1) # 옮길 거 이동할 영역을 정의하고 얼마나 움직일 지

# ws["B1"].value = "국어" #B1 셀어 '국어'입력

ws.move_range("C1:C11", rows=5, cols = -1)

wb.save("sample_korean.xlsx")



