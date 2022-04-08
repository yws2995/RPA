from openpyxl import Workbook
wb = Workbook()
# wb.active
ws = wb.create_sheet() # 기본으로 새로운 시트 생성
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RBG 형태로 값을 넣어 탭 색상 변경

ws1= wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # index 2부여	

new_ws = wb["NewSheet"] # Dict 형태로 시트에 접근

print(wb.sheetnames)

#Sheet 복시
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")